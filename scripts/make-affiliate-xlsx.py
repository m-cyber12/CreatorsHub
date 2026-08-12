#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
make-affiliate-xlsx.py (2026-08-10)
Builds DirectoryAI-Hub-Affiliate-Links.xlsx from the tool catalog dump.

Steps:
  1) npx tsx scripts/dump-tools.ts > /tmp/tools.json
  2) python3 scripts/make-affiliate-xlsx.py /tmp/tools.json

The sheet is sorted for the user's affiliate workflow:
  * PartnerStack-supported tools first (they already run programs the user can
    join), then everything else — both by rough revenue potential (higher
    starting price first, since there is no real revenue data on the site).
  * The yellow "لینک افیلیت شما" column is empty for the user to fill in.
  * A "شبکه" column shows the network each tool is known to use, so the user
    can filter (PartnerStack / Impact / Rewardful / FirstPromoter / direct).
"""
import json
import re
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── tools known to run their affiliate program on PartnerStack (verified
#    via public sources, 2026-08-10). Keys are catalog slugs.
PARTNERSTACK = {
    'descript': 'PartnerStack',
    'elevenlabs': 'PartnerStack',
    'murf-ai': 'PartnerStack',
    'synthesia': 'PartnerStack',
}
# other known networks (best-effort from public sources — editable)
KNOWN_NETWORK = {
    'invideo': 'Impact',
    'veed': 'Impact',
    'canva': 'Impact',
    'heygen': 'Rewardful',
    'pictory': 'FirstPromoter',
    'opusclip': 'Impact',
    'jasper': 'Impact',
    'runway': 'In-app',
}

SITE = 'https://creatorsaicenter.vercel.app'


def price_num(s: str) -> float:
    if not s:
        return 0.0
    m = re.search(r'([\d]+(?:\.[\d]+)?)', s.replace(',', '.'))
    return float(m.group(1)) if m else 0.0


def potential(t) -> int:
    price = price_num(t['startingPrice'])
    if t['pricing'] == 'Free':
        return 1
    if price == 0:
        return 2
    if price < 10:
        return 2
    if price < 30:
        return 3
    if price < 60:
        return 4
    return 5


def network_of(t) -> str:
    return PARTNERSTACK.get(t['slug']) or KNOWN_NETWORK.get(t['slug']) or ''


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else '/tmp/tools.json'
    tools = json.load(open(path, encoding='utf-8'))

    for t in tools:
        t['network'] = network_of(t)
        t['potential'] = potential(t)

    # PartnerStack first, then by potential (price), then name.
    tools.sort(key=lambda t: (
        0 if t['network'] == 'PartnerStack' else 1,
        -t['potential'],
        t['name'].lower(),
    ))

    wb = Workbook()

    # ── Sheet 1: tools ────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'ابزارها'

    header_fill = PatternFill('solid', fgColor='1F2937')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    ps_fill = PatternFill('solid', fgColor='DCFCE7')      # light green = PartnerStack
    aff_fill = PatternFill('solid', fgColor='FEF3C7')     # light amber = your link
    thin = Side(style='thin', color='D1D5DB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    cols = [
        ('ردیف', 6),
        ('نام ابزار', 30),
        ('اسلاگ', 22),
        ('دسته', 20),
        ('قیمت‌گذاری', 12),
        ('قیمت شروع', 12),
        ('شبکه', 14),
        ('وب‌سایت', 32),
        ('لینک فعلی (go)', 38),
        ('لینک افیلیت شما — اینجا بگذار', 44),
        ('پتانسیل ۱–۵', 10),
    ]
    for i, (title, width) in enumerate(cols, 1):
        c = ws.cell(row=1, column=i, value=title)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 22

    for idx, t in enumerate(tools, 1):
        r = idx + 1
        values = [
            idx,
            t['name'],
            t['slug'],
            t['category'],
            t['pricing'],
            t['startingPrice'],
            t['network'],
            t['url'],
            f'{SITE}/go/{t["slug"]}',
            '',  # user's affiliate link
            f"{t['potential']}/5",
        ]
        is_ps = t['network'] == 'PartnerStack'
        for i, v in enumerate(values, 1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = border
            if i == 10:
                c.fill = aff_fill
            elif is_ps:
                c.fill = ps_fill
            c.alignment = Alignment(vertical='center', wrap_text=(i in (8, 9, 10)))

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:K{len(tools) + 1}'

    # ── Sheet 2: guide ────────────────────────────────────────────────
    ws2 = wb.create_sheet('راهنما')
    guide = [
        ['راهنمای لینک‌های افیلیت — CreatorAI Hub'],
        [''],
        ['بخش افیلیت سایت:', ''],
        ['  صفحهٔ پیشنهادها و تخفیف‌ها (Deals):', f'{SITE}/deals'],
        ['  صفحهٔ افشای افیلیت:', f'{SITE}/disclosure'],
        ['  الگوی لینک هر ابزار:', f'{SITE}/go/{'{slug}'}'],
        [''],
        ['ابزارهای سبز = روی PartnerStack هستند (بالای لیست).', ''],
        ['  تأییدشده: Descript · ElevenLabs · Murf.ai · Synthesia', ''],
        ['  بقیه شبکه‌ها (Impact/Rewardful/…) در ستون «شبکه» آمده؛ قبل از ثبت، تأیید کن.', ''],
        [''],
        ['چطور کار می‌کند:', ''],
        ['  ۱) لینک افیلیت را از شبکه بگیر (PartnerStack / Impact / Rewardful / مستقیم).'],
        ['  ۲) در پنل ادمین → ابزارها → Edit → بخش «Affiliate link»: URL + انتخاب شبکه.'],
        ['  ۳) Save — تا چند ثانیه بعد، دکمه‌های Visit سایت به لینک افیلیت هدایت می‌کنند.'],
        ['  ۴) همین فایل را هم پر کن (ستون زرد) تا همه‌چیز یکجا ثبت باشد.'],
        [''],
        ['نکته: سایت فقط وقتی لینک افیلیت می‌فرستد که «شبکهٔ افیلیت» انتخاب شده باشد (قانون صداقت).'],
    ]
    for i, row in enumerate(guide, 1):
        for j, v in enumerate(row, 1):
            c = ws2.cell(row=i, column=j, value=v)
            if i == 1:
                c.font = Font(bold=True, size=14)
            elif j == 2 and v.startswith(SITE):
                c.font = Font(bold=True, color='2563EB')
            c.alignment = Alignment(vertical='top')
    ws2.column_dimensions['A'].width = 52
    ws2.column_dimensions['B'].width = 80

    out = '/home/user/DirectoryAI-Hub-Affiliate-Links.xlsx'
    wb.save(out)
    ps_count = sum(1 for t in tools if t['network'] == 'PartnerStack')
    print(f'saved {out} | tools: {len(tools)} | PartnerStack: {ps_count}')


if __name__ == '__main__':
    main()
